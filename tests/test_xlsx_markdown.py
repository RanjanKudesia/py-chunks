import pytest
import openpyxl
from py_chunks import get_markdown
from py_chunks.chunkers.xlsx import xlsx_to_markdown


def _make_xlsx(tmp_path, sheets):
    """sheets: list of (sheet_name, list_of_rows) where each row is a list of values."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets:
        ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)
    path = tmp_path / "test.xlsx"
    wb.save(str(path))
    return path


# ── Basic structure ───────────────────────────────────────────────────────────

def test_single_sheet_heading(tmp_path):
    path = _make_xlsx(tmp_path, [("Sales", [["A", "B"], ["1", "2"]])])
    result = xlsx_to_markdown(str(path))
    assert "## Sales" in result


def test_pipe_table_rendered(tmp_path):
    path = _make_xlsx(tmp_path, [("Data", [["Name", "Value"], ["Alpha", "42"]])])
    result = xlsx_to_markdown(str(path))
    assert "|" in result
    assert "Name" in result
    assert "Alpha" in result


def test_header_separator_present(tmp_path):
    path = _make_xlsx(tmp_path, [("Data", [["H1", "H2"], ["r1", "r2"], ["r3", "r4"]])])
    result = xlsx_to_markdown(str(path))
    assert "---" in result


def test_multi_sheet_both_present(tmp_path):
    path = _make_xlsx(tmp_path, [
        ("Sheet1", [["A", "B"], ["1", "2"]]),
        ("Sheet2", [["X", "Y"], ["9", "8"]]),
    ])
    result = xlsx_to_markdown(str(path))
    assert "## Sheet1" in result
    assert "## Sheet2" in result
    assert "Sheet1" in result
    assert "Sheet2" in result


def test_multi_sheet_separated_by_hr(tmp_path):
    path = _make_xlsx(tmp_path, [
        ("A", [["h"], ["v"]]),
        ("B", [["h"], ["v"]]),
    ])
    result = xlsx_to_markdown(str(path))
    assert "---" in result


def test_empty_sheet_skipped(tmp_path):
    path = _make_xlsx(tmp_path, [
        ("Empty", []),
        ("Full", [["X"], ["1"]]),
    ])
    result = xlsx_to_markdown(str(path))
    assert "## Full" in result
    assert "## Empty" not in result


def test_numeric_cells_rendered(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nums"
    ws.append(["Item", "Count"])
    ws.append(["Widget", 42])
    ws.append(["Gadget", 7])
    path = tmp_path / "nums.xlsx"
    wb.save(str(path))
    result = xlsx_to_markdown(str(path))
    assert "42" in result
    assert "7" in result


def test_pipe_characters_escaped(tmp_path):
    path = _make_xlsx(tmp_path, [("T", [["A|B", "C"], ["x|y", "z"]])])
    result = xlsx_to_markdown(str(path))
    assert "\\|" in result


# ── Error cases ───────────────────────────────────────────────────────────────

def test_missing_file_raises(tmp_path):
    with pytest.raises((FileNotFoundError, Exception)):
        xlsx_to_markdown(str(tmp_path / "no.xlsx"))


def test_wrong_extension_raises(tmp_path):
    path = tmp_path / "f.csv"
    path.write_bytes(b"a,b")
    with pytest.raises((ValueError, Exception)):
        xlsx_to_markdown(str(path))


def test_get_markdown_routes_xlsx(tmp_path):
    path = _make_xlsx(tmp_path, [("S", [["Col"], ["Val"]])])
    result = get_markdown(str(path))
    assert "Col" in result
    assert "Val" in result
