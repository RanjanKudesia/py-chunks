"""Focused contract tests for XLSX row-mode chunking."""

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from py_chunks import get_chunks, get_chunks_from_bytes, get_chunks_from_fileobj, stream_chunks
from py_chunks.chunkers.xlsx import chunk_xlsx, stream_chunk_xlsx


STANDARD_KEYS = {"content", "content_type", "metadata"}
METADATA_KEYS = {
    "sheet_name",
    "sheet_index",
    "row_index",
    "header_row",
    "col_count",
    "rows_per_chunk",
    "actual_row_count",
    "chunk_index",
}


def _workbook_bytes(*, with_second_sheet: bool = True) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append([None, None, None])
    ws.append(["Revenue", "Units", "Quarter"])
    ws.append([42000, 18, "Q3"])
    ws.append([None, None, None])
    ws.append([38000, 15, "Q4"])
    ws.append([12500, "", "Q1"])

    if with_second_sheet:
        ws2 = wb.create_sheet("Region")
        ws2.append(["Region", "Manager"])
        ws2.append(["West", "Avery"])
        ws2.append(["East", "Jordan"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_xlsx(tmp_path: Path, name: str = "sample.xlsx", **kwargs) -> Path:
    path = tmp_path / name
    path.write_bytes(_workbook_bytes(**kwargs))
    return path


def _assert_schema(chunks: list[dict]) -> None:
    assert chunks
    for chunk in chunks:
        assert STANDARD_KEYS.issubset(chunk.keys())
        assert chunk["content_type"] == "row_document"
        assert isinstance(chunk["content"], str) and chunk["content"]
        assert METADATA_KEYS.issubset(chunk["metadata"].keys())


def test_chunk_xlsx_schema_and_timing(tmp_path: Path):
    path = _write_xlsx(tmp_path)

    chunks, timing = chunk_xlsx(str(path))

    _assert_schema(chunks)
    assert timing["rust_ms"] > 0
    assert timing["python_ms"] > 0


def test_header_detection_and_empty_row_skipping(tmp_path: Path):
    path = _write_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), sheet_names=["Summary"])

    assert len(chunks) == 3
    assert chunks[0]["metadata"]["header_row"] == ["Revenue", "Units", "Quarter"]
    assert chunks[0]["metadata"]["row_index"] == 2
    assert chunks[0]["content"] == "Revenue: 42000 | Units: 18 | Quarter: Q3"
    assert chunks[1]["metadata"]["row_index"] == 4
    assert chunks[2]["content"] == "Revenue: 12500 | Units:  | Quarter: Q1"


def test_rows_per_chunk_groups_rows_and_chunk_index_is_sequential(tmp_path: Path):
    path = _write_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), sheet_names=["Summary"], rows_per_chunk=2)

    assert len(chunks) == 2
    assert chunks[0]["metadata"]["rows_per_chunk"] == 2
    assert chunks[0]["metadata"]["actual_row_count"] == 2
    assert chunks[1]["metadata"]["actual_row_count"] == 1
    assert chunks[0]["metadata"]["chunk_index"] == 0
    assert chunks[1]["metadata"]["chunk_index"] == 1
    assert "\n" in chunks[0]["content"]


def test_multi_sheet_default_and_filtering(tmp_path: Path):
    path = _write_xlsx(tmp_path)

    all_chunks, _ = chunk_xlsx(str(path))
    summary_only, _ = chunk_xlsx(str(path), sheet_names=["Summary"])
    region_only, _ = chunk_xlsx(str(path), sheet_names=["Region"])

    assert len(all_chunks) == len(summary_only) + len(region_only)
    assert {chunk["metadata"]["sheet_name"] for chunk in all_chunks} == {"Summary", "Region"}
    assert all(chunk["metadata"]["sheet_name"] == "Region" for chunk in region_only)


def test_chunk_index_restarts_per_sheet(tmp_path: Path):
    path = _write_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path))
    summary_indices = [c["metadata"]["chunk_index"] for c in chunks if c["metadata"]["sheet_name"] == "Summary"]
    region_indices = [c["metadata"]["chunk_index"] for c in chunks if c["metadata"]["sheet_name"] == "Region"]

    assert summary_indices == list(range(len(summary_indices)))
    assert region_indices == list(range(len(region_indices)))


def test_unified_get_chunks_maps_sentences_per_chunk_to_rows(tmp_path: Path):
    path = _write_xlsx(tmp_path)

    chunks = get_chunks(str(path), sentences_per_chunk=2)

    assert len(chunks) == 3
    assert chunks[0]["metadata"]["rows_per_chunk"] == 2
    assert chunks[0]["metadata"]["actual_row_count"] == 2


def test_stream_chunks_matches_batch_output(tmp_path: Path):
    path = _write_xlsx(tmp_path)

    batch_chunks, _ = chunk_xlsx(str(path), rows_per_chunk=2)
    streamed_chunks = list(stream_chunk_xlsx(str(path), rows_per_chunk=2))
    unified_streamed = list(stream_chunks(str(path), sentences_per_chunk=2))

    assert streamed_chunks == batch_chunks
    assert unified_streamed == batch_chunks


def test_get_chunks_from_bytes_and_fileobj():
    data = _workbook_bytes()

    byte_chunks = get_chunks_from_bytes(data, "report.xlsx", sentences_per_chunk=2)
    fileobj_chunks = get_chunks_from_fileobj(BytesIO(data), filename="report.xlsx", sentences_per_chunk=2)

    assert byte_chunks == fileobj_chunks
    assert byte_chunks[0]["metadata"]["rows_per_chunk"] == 2


def test_include_headers_false_uses_plain_row_serialization(tmp_path: Path):
    path = _write_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), sheet_names=["Summary"], include_headers=False)

    assert chunks[0]["content"] == "42000 | 18 | Q3"
    assert chunks[0]["metadata"]["header_row"] == ["Revenue", "Units", "Quarter"]


def test_error_cases(tmp_path: Path):
    path = tmp_path / "not_xlsx.txt"
    path.write_text("hello", encoding="utf-8")

    with pytest.raises(FileNotFoundError):
        chunk_xlsx(str(tmp_path / "missing.xlsx"))

    with pytest.raises(ValueError):
        chunk_xlsx(str(path))

    good_path = _write_xlsx(tmp_path)

    with pytest.raises(ValueError):
        chunk_xlsx(str(good_path), rows_per_chunk=0)

    with pytest.raises(ValueError, match="Sheet 'Unknown' not found"):
        chunk_xlsx(str(good_path), sheet_names=["Unknown"])


def test_get_chunks_default_mode_works_for_xlsx(tmp_path: Path):
    path = _write_xlsx(tmp_path)

    chunks = get_chunks(str(path))

    assert len(chunks) == 5
    assert all(chunk["content_type"] == "row_document" for chunk in chunks)


def _named_table_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Revenue", "Units", "Quarter"])
    ws.append([42000, 18, "Q3"])
    ws.append([38000, 15, "Q4"])
    ws.append([12500, 9, "Q1"])

    table = Table(displayName="SalesData", ref="A1:C4")
    ws.add_table(table)

    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["Region", "Manager"])
    ws2.append(["West", "Avery"])
    ws2.append(["East", "Jordan"])
    table2 = Table(displayName="RegionData", ref="A1:B3")
    ws2.add_table(table2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _heuristic_regions_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B", "C"])
    ws.append(["a1", "b1", "c1"])
    ws.append(["a2", "b2", "c2"])
    ws.append([None, None, None])
    ws.append(["X", "Y"])
    ws.append(["x1", "y1"])
    ws.append(["x2", "y2"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_named_table_xlsx(tmp_path: Path, name: str = "named.xlsx") -> Path:
    path = tmp_path / name
    path.write_bytes(_named_table_workbook_bytes())
    return path


def _write_heuristic_xlsx(tmp_path: Path, name: str = "heuristic.xlsx") -> Path:
    path = tmp_path / name
    path.write_bytes(_heuristic_regions_workbook_bytes())
    return path


TABLE_METADATA_KEYS = {
    "sheet_name",
    "sheet_index",
    "table_name",
    "is_named_table",
    "header_row",
    "start_row",
    "end_row",
    "start_col",
    "end_col",
    "row_count",
    "col_count",
    "chunk_index",
    "is_split",
    "split_part",
}


def test_table_mode_named_table_detection(tmp_path: Path):
    path = _write_named_table_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="table", sheet_names=["Sheet1"])

    assert len(chunks) == 1
    chunk = chunks[0]
    assert chunk["content_type"] == "table_region"
    assert chunk["metadata"]["is_named_table"] is True
    assert chunk["metadata"]["table_name"] == "SalesData"
    assert chunk["metadata"]["header_row"] == ["Revenue", "Units", "Quarter"]


def test_table_mode_named_table_range_metadata(tmp_path: Path):
    path = _write_named_table_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="table", sheet_names=["Sheet1"])

    metadata = chunks[0]["metadata"]
    assert metadata["start_row"] == 1
    assert metadata["end_row"] == 3
    assert metadata["start_col"] == 0
    assert metadata["end_col"] == 2


def test_table_mode_heuristic_regions(tmp_path: Path):
    path = _write_heuristic_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="table")

    assert len(chunks) == 2
    assert all(chunk["metadata"]["is_named_table"] is False for chunk in chunks)
    assert chunks[0]["metadata"]["header_row"] == ["A", "B", "C"]
    assert chunks[1]["metadata"]["header_row"] == ["X", "Y"]


def test_table_mode_multi_region_sheet_chunk_count(tmp_path: Path):
    path = _write_heuristic_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="table", sheet_names=["Sheet1"])

    assert len(chunks) == 2


def test_table_mode_chunk_splitting(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Revenue", "Units", "Quarter"])
    for i in range(1, 35):
        ws.append([10000 + i, i, f"Quarter-{i}"])
    table = Table(displayName="BigSales", ref="A1:C35")
    ws.add_table(table)

    path = tmp_path / "split.xlsx"
    buf = BytesIO()
    wb.save(buf)
    path.write_bytes(buf.getvalue())

    chunks, _ = chunk_xlsx(
        str(path),
        mode="table",
        max_chunk_chars=120,
        sheet_names=["Sheet1"],
    )

    assert len(chunks) > 1
    assert all(chunk["metadata"]["is_split"] is True for chunk in chunks)
    assert [chunk["metadata"]["split_part"] for chunk in chunks] == list(range(len(chunks)))


def test_table_mode_include_headers_false(tmp_path: Path):
    path = _write_named_table_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="table", include_headers=False, sheet_names=["Sheet1"])

    assert "Revenue:" not in chunks[0]["content"]
    assert "42000 | 18 | Q3" in chunks[0]["content"]
    assert chunks[0]["metadata"]["header_row"] == ["Revenue", "Units", "Quarter"]


def test_table_mode_sheet_filtering_and_multi_sheet(tmp_path: Path):
    path = _write_named_table_xlsx(tmp_path)

    all_chunks, _ = chunk_xlsx(str(path), mode="table")
    sheet1_chunks, _ = chunk_xlsx(str(path), mode="table", sheet_names=["Sheet1"])

    assert len(all_chunks) == 2
    assert len(sheet1_chunks) == 1
    assert all(chunk["metadata"]["sheet_name"] == "Sheet1" for chunk in sheet1_chunks)


def test_table_mode_stream_matches_batch(tmp_path: Path):
    path = _write_named_table_xlsx(tmp_path)

    batch_chunks, _ = chunk_xlsx(str(path), mode="table")
    streamed_chunks = list(stream_chunk_xlsx(str(path), mode="table"))

    assert streamed_chunks == batch_chunks


def test_table_mode_schema_and_chunk_index(tmp_path: Path):
    path = _write_named_table_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="table")

    for chunk in chunks:
        assert STANDARD_KEYS.issubset(chunk.keys())
        assert chunk["content_type"] == "table_region"
        assert TABLE_METADATA_KEYS.issubset(chunk["metadata"].keys())

    by_sheet = {}
    for chunk in chunks:
        by_sheet.setdefault(chunk["metadata"]["sheet_name"], []).append(chunk["metadata"]["chunk_index"])
    for indices in by_sheet.values():
        assert indices == list(range(len(indices)))


def test_table_mode_unknown_sheet_error(tmp_path: Path):
    path = _write_named_table_xlsx(tmp_path)

    with pytest.raises(ValueError, match="Sheet 'Missing' not found"):
        chunk_xlsx(str(path), mode="table", sheet_names=["Missing"])


def test_table_mode_unified_get_chunks(tmp_path: Path):
    path = _write_named_table_xlsx(tmp_path)

    chunks = get_chunks(str(path), mode="table")

    assert len(chunks) == 2
    assert all(chunk["content_type"] == "table_region" for chunk in chunks)


SHEET_METADATA_KEYS = {
    "sheet_name",
    "sheet_index",
    "row_count",
    "col_count",
    "header_row",
    "has_named_tables",
    "named_tables",
    "chunk_index",
    "is_split",
    "split_part",
}


def _sheet_mode_workbook_bytes(*, include_blank_sheet: bool = False, include_named_table: bool = True) -> bytes:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["Revenue", "Units", "Quarter"])
    ws1.append([42000, 18, "Q3"])
    ws1.append([38000, 15, "Q4"])
    if include_named_table:
        ws1.add_table(Table(displayName="SalesSheetTable", ref="A1:C3"))

    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["Region", "Manager", "Flag"])
    ws2.append(["West", "Avery"])
    ws2.append(["East", "Jordan", "Y"])

    if include_blank_sheet:
        wb.create_sheet("Blank")

    ws3 = wb.create_sheet("Sheet3")
    ws3.append(["Product", "Amount"])
    ws3.append(["Widget", 10])
    ws3.append(["Gadget", 12])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_sheet_mode_xlsx(tmp_path: Path, name: str = "sheet_mode.xlsx", **kwargs) -> Path:
    path = tmp_path / name
    path.write_bytes(_sheet_mode_workbook_bytes(**kwargs))
    return path


def test_sheet_mode_schema(tmp_path: Path):
    path = _write_sheet_mode_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="sheet")

    assert chunks
    for chunk in chunks:
        assert STANDARD_KEYS.issubset(chunk.keys())
        assert chunk["content_type"] == "sheet"
        assert SHEET_METADATA_KEYS.issubset(chunk["metadata"].keys())


def test_sheet_mode_one_chunk_per_non_empty_sheet(tmp_path: Path):
    path = _write_sheet_mode_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="sheet")

    assert len(chunks) == 3


def test_sheet_mode_empty_sheet_skipped_and_chunk_index_sequential(tmp_path: Path):
    path = _write_sheet_mode_xlsx(tmp_path, include_blank_sheet=True)

    chunks, _ = chunk_xlsx(str(path), mode="sheet")

    assert len(chunks) == 3
    assert [chunk["metadata"]["chunk_index"] for chunk in chunks] == [0, 1, 2]


def test_sheet_mode_row_col_counts(tmp_path: Path):
    path = _write_sheet_mode_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="sheet", sheet_names=["Sheet2"])

    assert len(chunks) == 1
    metadata = chunks[0]["metadata"]
    assert metadata["row_count"] == 2
    assert metadata["col_count"] == 3


def test_sheet_mode_named_tables_detected(tmp_path: Path):
    path = _write_sheet_mode_xlsx(tmp_path, include_named_table=True)

    chunks, _ = chunk_xlsx(str(path), mode="sheet", sheet_names=["Sheet1"])

    assert chunks[0]["metadata"]["has_named_tables"] is True
    assert "SalesSheetTable" in chunks[0]["metadata"]["named_tables"]


def test_sheet_mode_no_named_tables(tmp_path: Path):
    path = _write_sheet_mode_xlsx(tmp_path, include_named_table=False)

    chunks, _ = chunk_xlsx(str(path), mode="sheet", sheet_names=["Sheet1"])

    assert chunks[0]["metadata"]["has_named_tables"] is False
    assert chunks[0]["metadata"]["named_tables"] == []


def test_sheet_mode_sheet_filtering(tmp_path: Path):
    path = _write_sheet_mode_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="sheet", sheet_names=["Sheet1"])

    assert len(chunks) == 1
    assert chunks[0]["metadata"]["sheet_name"] == "Sheet1"


def test_sheet_mode_include_headers_false(tmp_path: Path):
    path = _write_sheet_mode_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="sheet", include_headers=False, sheet_names=["Sheet1"])

    assert "Revenue:" not in chunks[0]["content"]
    assert "42000 | 18 | Q3" in chunks[0]["content"]
    assert chunks[0]["metadata"]["header_row"] == ["Revenue", "Units", "Quarter"]


def test_sheet_mode_stream_matches_batch(tmp_path: Path):
    path = _write_sheet_mode_xlsx(tmp_path)

    batch_chunks, _ = chunk_xlsx(str(path), mode="sheet")
    streamed_chunks = list(stream_chunk_xlsx(str(path), mode="sheet"))

    assert streamed_chunks == batch_chunks


def test_sheet_mode_unified_get_chunks(tmp_path: Path):
    path = _write_sheet_mode_xlsx(tmp_path)

    chunks = get_chunks(str(path), mode="sheet")

    assert len(chunks) == 3
    assert all(chunk["content_type"] == "sheet" for chunk in chunks)


def test_sheet_mode_unknown_sheet_error(tmp_path: Path):
    path = _write_sheet_mode_xlsx(tmp_path)

    with pytest.raises(ValueError, match="Sheet 'Missing' not found"):
        chunk_xlsx(str(path), mode="sheet", sheet_names=["Missing"])


SLIDING_METADATA_KEYS = {
    "sheet_name",
    "sheet_index",
    "window_size",
    "overlap",
    "actual_row_count",
    "window_index",
    "start_row",
    "end_row",
    "header_row",
    "col_count",
    "chunk_index",
}


def _sliding_window_workbook_bytes(*, with_second_sheet: bool = False) -> bytes:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["Revenue", "Units", "Quarter"])
    ws1.append([42000, 18, "Q3"])
    ws1.append([38000, 15, "Q4"])
    ws1.append([12500, 6, "Q1"])
    ws1.append([51000, 20, "Q2"])
    ws1.append([47000, 17, "Q3"])
    ws1.append([39000, 11, "Q4"])

    if with_second_sheet:
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["Region", "Manager"])
        ws2.append(["West", "Avery"])
        ws2.append(["East", "Jordan"])
        ws2.append(["North", "Taylor"])
        ws2.append(["South", "Parker"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_sliding_window_xlsx(tmp_path: Path, name: str = "sliding.xlsx", **kwargs) -> Path:
    path = tmp_path / name
    path.write_bytes(_sliding_window_workbook_bytes(**kwargs))
    return path


def test_sliding_window_schema(tmp_path: Path):
    path = _write_sliding_window_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="sliding_window", window_size=3, overlap=1)

    assert chunks
    for chunk in chunks:
        assert STANDARD_KEYS.issubset(chunk.keys())
        assert chunk["content_type"] == "row_window"
        assert SLIDING_METADATA_KEYS.issubset(chunk["metadata"].keys())


def test_sliding_window_count_and_last_window_size(tmp_path: Path):
    path = _write_sliding_window_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="sliding_window", window_size=3, overlap=1)

    assert len(chunks) == 3
    assert [chunk["metadata"]["actual_row_count"] for chunk in chunks] == [3, 3, 2]
    assert chunks[-1]["metadata"]["actual_row_count"] <= 3


def test_sliding_window_overlap_shares_rows(tmp_path: Path):
    path = _write_sliding_window_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="sliding_window", window_size=3, overlap=1)

    lines0 = chunks[0]["content"].splitlines()
    lines1 = chunks[1]["content"].splitlines()
    lines2 = chunks[2]["content"].splitlines()
    assert lines0[-1] == lines1[0]
    assert lines1[-1] == lines2[0]


def test_sliding_window_size_one_overlap_zero_equals_row_mode(tmp_path: Path):
    path = _write_sliding_window_xlsx(tmp_path)

    sliding_chunks, _ = chunk_xlsx(
        str(path),
        mode="sliding_window",
        window_size=1,
        overlap=0,
    )
    row_chunks, _ = chunk_xlsx(str(path), mode="row", rows_per_chunk=1)

    assert len(sliding_chunks) == len(row_chunks)
    assert [c["content"] for c in sliding_chunks] == [c["content"] for c in row_chunks]


def test_sliding_window_window_index_resets_per_sheet_and_chunk_index_global(tmp_path: Path):
    path = _write_sliding_window_xlsx(tmp_path, with_second_sheet=True)

    chunks, _ = chunk_xlsx(str(path), mode="sliding_window", window_size=3, overlap=1)

    s1 = [c for c in chunks if c["metadata"]["sheet_name"] == "Sheet1"]
    s2 = [c for c in chunks if c["metadata"]["sheet_name"] == "Sheet2"]
    assert [c["metadata"]["window_index"] for c in s1] == [0, 1, 2]
    assert [c["metadata"]["window_index"] for c in s2] == [0, 1]
    assert [c["metadata"]["chunk_index"] for c in chunks] == list(range(len(chunks)))


def test_sliding_window_headers_prepended(tmp_path: Path):
    path = _write_sliding_window_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="sliding_window", window_size=3, overlap=1)

    assert all("Revenue:" in line for chunk in chunks for line in chunk["content"].splitlines())


def test_sliding_window_include_headers_false(tmp_path: Path):
    path = _write_sliding_window_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(
        str(path),
        mode="sliding_window",
        window_size=3,
        overlap=1,
        include_headers=False,
    )

    assert "Revenue:" not in chunks[0]["content"]
    assert "42000 | 18 | Q3" in chunks[0]["content"]
    assert chunks[0]["metadata"]["header_row"] == ["Revenue", "Units", "Quarter"]


def test_sliding_window_sheet_filtering(tmp_path: Path):
    path = _write_sliding_window_xlsx(tmp_path, with_second_sheet=True)

    chunks, _ = chunk_xlsx(
        str(path),
        mode="sliding_window",
        window_size=3,
        overlap=1,
        sheet_names=["Sheet1"],
    )

    assert len(chunks) == 3
    assert all(chunk["metadata"]["sheet_name"] == "Sheet1" for chunk in chunks)


def test_sliding_window_stream_matches_batch(tmp_path: Path):
    path = _write_sliding_window_xlsx(tmp_path)

    batch_chunks, _ = chunk_xlsx(str(path), mode="sliding_window", window_size=3, overlap=1)
    streamed_chunks = list(
        stream_chunk_xlsx(str(path), mode="sliding_window", window_size=3, overlap=1)
    )

    assert streamed_chunks == batch_chunks


def test_sliding_window_unified_get_chunks(tmp_path: Path):
    path = _write_sliding_window_xlsx(tmp_path)

    chunks = get_chunks(str(path), mode="sliding_window", window_size=3, overlap=1)

    assert len(chunks) == 3
    assert all(chunk["content_type"] == "row_window" for chunk in chunks)


def test_sliding_window_errors(tmp_path: Path):
    path = _write_sliding_window_xlsx(tmp_path)

    with pytest.raises(ValueError):
        chunk_xlsx(str(path), mode="sliding_window", window_size=0, overlap=0)

    with pytest.raises(ValueError):
        chunk_xlsx(str(path), mode="sliding_window", window_size=2, overlap=2)

    with pytest.raises(ValueError, match="Sheet 'Missing' not found"):
        chunk_xlsx(
            str(path),
            mode="sliding_window",
            window_size=3,
            overlap=1,
            sheet_names=["Missing"],
        )


PAGE_AWARE_METADATA_KEYS = {
    "sheet_name",
    "sheet_index",
    "has_print_area",
    "print_area_ref",
    "start_row",
    "end_row",
    "start_col",
    "end_col",
    "row_count",
    "col_count",
    "header_row",
    "region_index",
    "chunk_index",
    "is_split",
    "split_part",
}


def _page_aware_workbook_bytes() -> bytes:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["Revenue", "Units", "Quarter", "Ignore"])
    ws1.append([42000, 18, "Q3", "X"])
    ws1.append([38000, 15, "Q4", "Y"])
    ws1.append([12500, 6, "Q1", "Z"])
    ws1.append([99999, 999, "OUT", "OUT"])
    ws1.print_area = "A1:C4"

    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["Region", "Manager", "Flag"])
    ws2.append(["West", "Avery", "W"])
    ws2.append(["East", "Jordan", "E"])
    # No print area for fallback test.

    ws3 = wb.create_sheet("Sheet3")
    ws3.append(["H1", "H2", "H3"])
    ws3.append(["a1", "b1", "c1"])
    ws3.append(["a2", "b2", "c2"])
    ws3.append(["a3", "b3", "c3"])
    ws3.append([None, None, None])
    ws3.append(["H1b", "H2b", "H3b"])
    ws3.append(["d1", "e1", "f1"])
    ws3.append(["d2", "e2", "f2"])
    ws3.append(["d3", "e3", "f3"])
    ws3.append(["d4", "e4", "f4"])
    ws3.print_area = "A1:C4,A6:C10"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_page_aware_xlsx(tmp_path: Path, name: str = "page_aware.xlsx") -> Path:
    path = tmp_path / name
    path.write_bytes(_page_aware_workbook_bytes())
    return path


def test_page_aware_schema(tmp_path: Path):
    path = _write_page_aware_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="page_aware")

    assert chunks
    for chunk in chunks:
        assert STANDARD_KEYS.issubset(chunk.keys())
        assert chunk["content_type"] == "sheet_region"
        assert PAGE_AWARE_METADATA_KEYS.issubset(chunk["metadata"].keys())


def test_page_aware_print_area_respected(tmp_path: Path):
    path = _write_page_aware_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="page_aware", sheet_names=["Sheet1"])

    assert len(chunks) == 1
    content = chunks[0]["content"]
    assert "Revenue: 42000 | Units: 18 | Quarter: Q3" in content
    assert "Revenue: 99999" not in content
    assert chunks[0]["metadata"]["print_area_ref"] == "A1:C4"


def test_page_aware_fallback_without_print_area(tmp_path: Path):
    path = _write_page_aware_xlsx(tmp_path)

    page_chunks, _ = chunk_xlsx(str(path), mode="page_aware", sheet_names=["Sheet2"])
    sheet_chunks, _ = chunk_xlsx(str(path), mode="sheet", sheet_names=["Sheet2"])

    assert len(page_chunks) == 1
    assert page_chunks[0]["metadata"]["has_print_area"] is False
    assert page_chunks[0]["metadata"]["print_area_ref"] is None
    assert page_chunks[0]["content"] == sheet_chunks[0]["content"]


def test_page_aware_multiple_print_areas_one_sheet(tmp_path: Path):
    path = _write_page_aware_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="page_aware", sheet_names=["Sheet3"])

    assert len(chunks) == 2
    assert [c["metadata"]["region_index"] for c in chunks] == [0, 1]


def test_page_aware_region_index_resets_chunk_index_global(tmp_path: Path):
    path = _write_page_aware_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="page_aware")

    by_sheet = {}
    for c in chunks:
        by_sheet.setdefault(c["metadata"]["sheet_name"], []).append(c["metadata"]["region_index"])
    for indices in by_sheet.values():
        assert indices == list(range(len(indices)))
    assert [c["metadata"]["chunk_index"] for c in chunks] == list(range(len(chunks)))


def test_page_aware_row_col_count_accuracy(tmp_path: Path):
    path = _write_page_aware_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="page_aware", sheet_names=["Sheet1"])
    metadata = chunks[0]["metadata"]
    assert metadata["row_count"] == 3
    assert metadata["col_count"] == 3


def test_page_aware_include_headers_false(tmp_path: Path):
    path = _write_page_aware_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(
        str(path),
        mode="page_aware",
        include_headers=False,
        sheet_names=["Sheet1"],
    )

    assert "Revenue:" not in chunks[0]["content"]
    assert "42000 | 18 | Q3" in chunks[0]["content"]
    assert chunks[0]["metadata"]["header_row"] == ["Revenue", "Units", "Quarter"]


def test_page_aware_sheet_filtering(tmp_path: Path):
    path = _write_page_aware_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="page_aware", sheet_names=["Sheet1"])

    assert len(chunks) == 1
    assert all(c["metadata"]["sheet_name"] == "Sheet1" for c in chunks)


def test_page_aware_stream_matches_batch(tmp_path: Path):
    path = _write_page_aware_xlsx(tmp_path)

    batch_chunks, _ = chunk_xlsx(str(path), mode="page_aware")
    streamed_chunks = list(stream_chunk_xlsx(str(path), mode="page_aware"))

    assert streamed_chunks == batch_chunks


def test_page_aware_unified_get_chunks(tmp_path: Path):
    path = _write_page_aware_xlsx(tmp_path)

    chunks = get_chunks(str(path), mode="page_aware")

    assert chunks
    assert all(c["content_type"] == "sheet_region" for c in chunks)


def test_page_aware_unknown_sheet_error(tmp_path: Path):
    path = _write_page_aware_xlsx(tmp_path)

    with pytest.raises(ValueError, match="Sheet 'Missing' not found"):
        chunk_xlsx(str(path), mode="page_aware", sheet_names=["Missing"])


SEMANTIC_METADATA_KEYS = {
    "sheet_name",
    "sheet_index",
    "category_column",
    "category_value",
    "used_fallback",
    "low_grouping_quality",
    "avg_group_size",
    "start_row",
    "end_row",
    "actual_row_count",
    "header_row",
    "col_count",
    "group_index",
    "chunk_index",
}


def _semantic_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["Region", "Product", "Revenue"])
    ws.append(["West", "Widget", 100])
    ws.append(["West", "Gadget", 200])
    ws.append(["East", "Widget", 150])
    ws.append(["East", "Gadget", 250])
    ws.append(["East", "Doohickey", 300])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_semantic_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "semantic.xlsx"
    path.write_bytes(_semantic_workbook_bytes())
    return path


def test_semantic_schema(tmp_path: Path):
    path = _write_semantic_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="semantic")

    assert chunks
    for chunk in chunks:
        assert STANDARD_KEYS.issubset(chunk.keys())
        assert chunk["content_type"] == "semantic_group"
        assert SEMANTIC_METADATA_KEYS.issubset(chunk["metadata"].keys())


def test_semantic_grouping(tmp_path: Path):
    path = _write_semantic_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="semantic")

    # Pre-sort by category value means East < West alphabetically.
    assert len(chunks) == 2
    assert chunks[0]["metadata"]["category_value"] == "East"
    assert chunks[0]["metadata"]["actual_row_count"] == 3
    assert chunks[1]["metadata"]["category_value"] == "West"
    assert chunks[1]["metadata"]["actual_row_count"] == 2
    assert chunks[0]["metadata"]["used_fallback"] is False
    assert chunks[1]["metadata"]["used_fallback"] is False


def test_semantic_group_index_resets_per_sheet(tmp_path: Path):
    wb = Workbook()
    s1 = wb.active
    s1.title = "S1"
    s1.append(["Region", "Value"])
    s1.append(["West", 1])
    s1.append(["West", 2])
    s1.append(["East", 3])

    s2 = wb.create_sheet("S2")
    s2.append(["Region", "Value"])
    s2.append(["North", 1])
    s2.append(["North", 2])
    s2.append(["South", 3])

    path = tmp_path / "semantic_multisheet.xlsx"
    buf = BytesIO()
    wb.save(buf)
    path.write_bytes(buf.getvalue())

    chunks, _ = chunk_xlsx(str(path), mode="semantic")

    s1_groups = [c["metadata"]["group_index"] for c in chunks if c["metadata"]["sheet_name"] == "S1"]
    s2_groups = [c["metadata"]["group_index"] for c in chunks if c["metadata"]["sheet_name"] == "S2"]
    assert s1_groups == [0, 1]
    assert s2_groups == [0, 1]
    assert [c["metadata"]["chunk_index"] for c in chunks] == list(range(len(chunks)))


def test_semantic_fallback_unique_values(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Unique"
    ws.append(["Code", "Amount"])
    ws.append(["A", 1])
    ws.append(["B", 2])
    ws.append(["C", 3])
    ws.append(["D", 4])
    ws.append(["E", 5])

    path = tmp_path / "semantic_unique.xlsx"
    buf = BytesIO()
    wb.save(buf)
    path.write_bytes(buf.getvalue())

    chunks, _ = chunk_xlsx(str(path), mode="semantic", rows_per_chunk=2)

    assert len(chunks) == 3
    assert all(c["metadata"]["used_fallback"] is True for c in chunks)
    assert all(c["metadata"]["category_column"] is None for c in chunks)
    assert all(c["metadata"]["category_value"] is None for c in chunks)
    assert [c["metadata"]["actual_row_count"] for c in chunks] == [2, 2, 1]


def test_semantic_streaming_parity(tmp_path: Path):
    path = _write_semantic_xlsx(tmp_path)

    batch_chunks, _ = chunk_xlsx(str(path), mode="semantic")
    streamed_chunks = list(stream_chunk_xlsx(str(path), mode="semantic"))

    assert streamed_chunks == batch_chunks


def test_semantic_unified_api(tmp_path: Path):
    path = _write_semantic_xlsx(tmp_path)

    chunks = get_chunks(str(path), mode="semantic")

    assert chunks
    assert all(chunk["content_type"] == "semantic_group" for chunk in chunks)


# ============================================================
# Quality tests against real xlsx files in test_files/excel/
# ============================================================

_REAL_EXCEL_DIR = Path(__file__).parent.parent.parent / "test_files" / "excel"
_REAL_FILES = {
    "10": _REAL_EXCEL_DIR / "file_example_XLSX_10.xlsx",
    "50": _REAL_EXCEL_DIR / "file_example_XLSX_50.xlsx",
    "100": _REAL_EXCEL_DIR / "file_example_XLSX_100.xlsx",
    "1000": _REAL_EXCEL_DIR / "file_example_XLSX_1000.xlsx",
    "5000": _REAL_EXCEL_DIR / "file_example_XLSX_5000.xlsx",
}
_real_files_available = all(f.exists() for f in _REAL_FILES.values())
real_files = pytest.mark.skipif(not _real_files_available, reason="test_files/excel/ not found")


@real_files
@pytest.mark.parametrize("size", ["10", "50", "100", "1000", "5000"])
def test_real_xlsx_all_modes_no_error(size):
    path = str(_REAL_FILES[size])
    for mode in ["row", "table", "sheet", "sliding_window", "page_aware", "semantic"]:
        kwargs = {"window_size": 5, "overlap": 2} if mode == "sliding_window" else {}
        chunks, timing = chunk_xlsx(path, mode=mode, **kwargs)
        assert isinstance(chunks, list)
        assert timing["rust_ms"] > 0


@real_files
@pytest.mark.parametrize("size,expected_rows", [
    ("10", 10),
    ("50", 50),
    ("100", 100),
    ("1000", 1000),
    ("5000", 5000),
])
def test_real_xlsx_row_mode_total_rows(size, expected_rows):
    """Total rows across all chunks should match the file's row count (±1 for ambiguous header rows)."""
    path = str(_REAL_FILES[size])
    chunks, _ = chunk_xlsx(path, mode="row", rows_per_chunk=5)
    total = sum(c["metadata"]["actual_row_count"] for c in chunks)
    assert abs(total - expected_rows) <= 1


@real_files
@pytest.mark.parametrize("size", ["1000", "5000"])
def test_real_xlsx_large_files_header_detection(size):
    """1000 and 5000-row files have proper string headers that should be detected."""
    path = str(_REAL_FILES[size])
    chunks, _ = chunk_xlsx(path, mode="row", rows_per_chunk=1)
    header = chunks[0]["metadata"]["header_row"]
    named = [h for h in header if not h.startswith("Column ")]
    assert len(named) >= 3, f"Expected named headers, got: {header}"


@real_files
@pytest.mark.parametrize("size", ["1000", "5000"])
def test_real_xlsx_semantic_picks_gender_column(size):
    """After the cardinality fix, semantic mode picks Gender (2 values) over First Name (50 values)."""
    path = str(_REAL_FILES[size])
    chunks, _ = chunk_xlsx(path, mode="semantic")
    non_fallback = [c for c in chunks if not c["metadata"]["used_fallback"]]
    assert non_fallback, "Expected category-based grouping, got only fallback chunks"
    cat_vals = {c["metadata"]["category_value"] for c in non_fallback}
    assert cat_vals == {"Male", "Female"}, f"Expected Gender grouping, got: {cat_vals}"


@real_files
@pytest.mark.parametrize("size", ["1000", "5000"])
def test_real_xlsx_semantic_fewer_chunks_than_rows(size):
    """Gender grouping collapses consecutive same-gender runs so chunks < total rows."""
    path = str(_REAL_FILES[size])
    chunks, _ = chunk_xlsx(path, mode="semantic")
    expected_rows = int(size)
    assert len(chunks) < expected_rows, (
        f"Expected fewer chunks than rows for {size}-row file, got {len(chunks)}"
    )


@real_files
def test_real_xlsx_sheet_mode_all_files():
    """Sheet mode produces exactly 1 chunk per file when content fits within the limit."""
    for size, path in _REAL_FILES.items():
        chunks, _ = chunk_xlsx(str(path), mode="sheet", max_chunk_chars=10_000_000)
        assert len(chunks) == 1, f"Expected 1 sheet chunk for {size}-row file, got {len(chunks)}"
        assert chunks[0]["content_type"] == "sheet"
        assert chunks[0]["metadata"]["row_count"] > 0


@real_files
def test_real_xlsx_table_mode_finds_regions():
    """Table mode finds at least 1 region in every real file."""
    for size, path in _REAL_FILES.items():
        chunks, _ = chunk_xlsx(str(path), mode="table")
        assert chunks, f"table mode returned no chunks for {size}-row file"
        assert all(c["content_type"] in ("table", "table_region") for c in chunks)


@real_files
def test_real_xlsx_sliding_window_metadata_correct():
    """Sliding window metadata is self-consistent: actual_row_count ≤ window_size, overlap correct."""
    path = str(_REAL_FILES["1000"])
    w, o = 10, 3
    chunks, _ = chunk_xlsx(path, mode="sliding_window", window_size=w, overlap=o)
    assert all(c["metadata"]["window_size"] == w for c in chunks)
    assert all(c["metadata"]["overlap"] == o for c in chunks)
    assert all(c["metadata"]["actual_row_count"] <= w for c in chunks)
    assert chunks[-1]["metadata"]["actual_row_count"] >= 1


@real_files
def test_real_xlsx_page_aware_fallback_for_files_without_print_area():
    """Real files have no print areas, so page_aware falls back to full-sheet chunks."""
    for size, path in _REAL_FILES.items():
        chunks, _ = chunk_xlsx(str(path), mode="page_aware", max_chunk_chars=10_000_000)
        assert len(chunks) == 1, f"Expected 1 page_aware chunk for {size}-row file"
        assert chunks[0]["metadata"]["has_print_area"] is False
        assert chunks[0]["metadata"]["print_area_ref"] is None


@real_files
@pytest.mark.parametrize("size", ["1000", "5000"])
def test_real_xlsx_chunk_index_globally_sequential(size):
    """chunk_index must be 0, 1, 2, ... across all chunks for every mode."""
    path = str(_REAL_FILES[size])
    for mode in ["row", "table", "sheet", "sliding_window", "page_aware", "semantic"]:
        kwargs = {"window_size": 5, "overlap": 2} if mode == "sliding_window" else {}
        chunks, _ = chunk_xlsx(path, mode=mode, **kwargs)
        indices = [c["metadata"]["chunk_index"] for c in chunks]
        assert indices == list(range(len(chunks))), f"Non-sequential chunk_index in {mode} mode"


# ============================================================
# Fix verification tests
# ============================================================

def test_header_detection_numeric_index_column(tmp_path: Path):
    """A numeric first cell (e.g. row index 0) followed by all-string cells is now
    treated as a header row, matching the behaviour of files that store a sequential
    integer index in column A."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append([0, "Name", "Score"])   # numeric index + string labels
    ws.append([1, "Alice", 90])
    ws.append([2, "Bob", 85])
    ws.append([3, "Carol", 95])

    path = tmp_path / "indexed.xlsx"
    buf = BytesIO()
    wb.save(buf)
    path.write_bytes(buf.getvalue())

    chunks, _ = chunk_xlsx(str(path), mode="row", rows_per_chunk=1)

    assert len(chunks) == 3, "Header row must not be treated as a data row"
    header = chunks[0]["metadata"]["header_row"]
    assert header[1] == "Name"
    assert header[2] == "Score"


def test_semantic_pre_sort_groups_alphabetically(tmp_path: Path):
    """After the pre-sort fix, chunks are emitted in category-value alphabetical order
    regardless of the original row order in the file."""
    path = _write_semantic_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="semantic")

    cat_values = [c["metadata"]["category_value"] for c in chunks]
    assert cat_values == sorted(set(cat_values)), "Chunks must be in alphabetical category order"


def test_semantic_low_grouping_quality_flag(tmp_path: Path):
    """low_grouping_quality is True when avg rows per group < 2 (many singleton groups)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sparse"
    ws.append(["Category", "Value"])
    ws.append(["A", 1])   # A appears twice → column has repetition, gets detected
    ws.append(["A", 2])
    for ch in "BCDEFGHI":  # 8 categories, each appearing once
        ws.append([ch, ord(ch)])

    path = tmp_path / "sparse.xlsx"
    buf = BytesIO()
    wb.save(buf)
    path.write_bytes(buf.getvalue())

    chunks, _ = chunk_xlsx(str(path), mode="semantic")

    non_fallback = [c for c in chunks if not c["metadata"]["used_fallback"]]
    assert non_fallback, "Expected category-based grouping"
    # 9 groups (A=2 rows, B…I each 1 row) → avg = 10/9 ≈ 1.11 < 2.0
    assert non_fallback[0]["metadata"]["low_grouping_quality"] is True
    assert non_fallback[0]["metadata"]["avg_group_size"] < 2.0


def test_semantic_good_grouping_quality(tmp_path: Path):
    """low_grouping_quality is False when avg rows per group >= 2."""
    path = _write_semantic_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="semantic")

    # East=3, West=2 → avg=2.5 ≥ 2.0
    assert all(not c["metadata"]["low_grouping_quality"] for c in chunks)
    assert all(c["metadata"]["avg_group_size"] >= 2.0 for c in chunks)


def test_sheet_mode_max_chunk_chars_splits(tmp_path: Path):
    """Sheet mode splits large content into multiple chunks when max_chunk_chars is exceeded."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Revenue", "Units", "Quarter"])
    for i in range(1, 21):
        ws.append([10000 + i, i, f"Quarter-{i}"])

    path = tmp_path / "big_sheet.xlsx"
    buf = BytesIO()
    wb.save(buf)
    path.write_bytes(buf.getvalue())

    chunks, _ = chunk_xlsx(str(path), mode="sheet", max_chunk_chars=200)

    assert len(chunks) > 1, "Expected content to be split across multiple chunks"
    for chunk in chunks:
        assert chunk["content_type"] == "sheet"
        assert chunk["metadata"]["is_split"] is True
        assert isinstance(chunk["metadata"]["split_part"], int)

    assert [c["metadata"]["split_part"] for c in chunks] == list(range(len(chunks)))
    total_rows = sum(c["metadata"]["row_count"] for c in chunks)
    assert total_rows == 20


def test_sheet_mode_no_split_below_limit(tmp_path: Path):
    """When content fits within max_chunk_chars, is_split is False and split_part is None."""
    path = _write_sheet_mode_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="sheet", max_chunk_chars=100_000)

    for chunk in chunks:
        assert chunk["metadata"]["is_split"] is False
        assert chunk["metadata"]["split_part"] is None


def test_page_aware_max_chunk_chars_splits(tmp_path: Path):
    """Page-aware mode splits large fallback content when max_chunk_chars is exceeded."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Revenue", "Units", "Quarter"])
    for i in range(1, 21):
        ws.append([10000 + i, i, f"Quarter-{i}"])

    path = tmp_path / "big_page.xlsx"
    buf = BytesIO()
    wb.save(buf)
    path.write_bytes(buf.getvalue())

    chunks, _ = chunk_xlsx(str(path), mode="page_aware", max_chunk_chars=200)

    assert len(chunks) > 1, "Expected content to be split across multiple chunks"
    for chunk in chunks:
        assert chunk["content_type"] == "sheet_region"
        assert chunk["metadata"]["is_split"] is True

    total_rows = sum(c["metadata"]["row_count"] for c in chunks)
    assert total_rows == 20


def test_page_aware_no_split_below_limit(tmp_path: Path):
    """When content fits within max_chunk_chars, is_split is False."""
    path = _write_page_aware_xlsx(tmp_path)

    chunks, _ = chunk_xlsx(str(path), mode="page_aware", max_chunk_chars=100_000)

    for chunk in chunks:
        assert chunk["metadata"]["is_split"] is False
        assert chunk["metadata"]["split_part"] is None


@real_files
@pytest.mark.parametrize("size", ["1000", "5000"])
def test_real_xlsx_semantic_pre_sort_reduces_chunks(size):
    """After pre-sort, Gender grouping yields exactly 2 consecutive runs (Female, Male)."""
    path = str(_REAL_FILES[size])
    chunks, _ = chunk_xlsx(path, mode="semantic")
    non_fallback = [c for c in chunks if not c["metadata"]["used_fallback"]]
    assert len(non_fallback) == 2, (
        f"Pre-sorted Gender should yield 2 chunks, got {len(non_fallback)}"
    )
    cat_values = [c["metadata"]["category_value"] for c in non_fallback]
    assert cat_values == sorted(cat_values), "Chunks must be in alphabetical order"


@real_files
@pytest.mark.parametrize("size", ["10", "50", "100"])
def test_real_xlsx_small_files_header_now_detected(size):
    """After the numeric-index-column header fix, small real files detect headers correctly."""
    path = str(_REAL_FILES[size])
    chunks, _ = chunk_xlsx(path, mode="row", rows_per_chunk=1)
    header = chunks[0]["metadata"]["header_row"]
    # The header row in these files is: 0, First Name, Last Name, Gender, ...
    named = [h for h in header if h not in ("", "0") and not h.startswith("Column ")]
    assert len(named) >= 3, f"Expected named headers after fix, got: {header}"
    assert "First Name" in header or any("Name" in h for h in header)